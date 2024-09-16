from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.http import JsonResponse
import tempfile
import pandas as pd
import sqlite3
import time
from multiprocessing import Process, Manager
from xlsxwriter.workbook import Workbook
import csv
import os
import shutil
import warnings
from datetime import datetime
from datetime import timedelta
from pathlib import Path
import sys
import zipfile
import calendar
import openpyxl
from .models import ProcessingTask


path_root = Path(__file__).parents[1] / "Scripts"
sys.path.append(str(path_root))

warnings.filterwarnings("ignore", category=FutureWarning) 
warnings.filterwarnings("ignore", category=DeprecationWarning)


conn_main = sqlite3.connect("Data/Databases/Data.sqlite3", check_same_thread=False)
cur_main = conn_main.cursor()

count_lis = []
df_dict = {}
process_dict = {}

def get_latest_date():
    latest_month = (cur_main.execute("select max(stored_months) from master").fetchall()[0][0]).split("_")
    latest_month.append(str(calendar.monthrange(int(latest_month[0]), int(latest_month[1]))[1]))
    return "-".join(latest_month)

def run_query(query, year, dict1):
    conn = sqlite3.connect(f"Data/Databases/Data.sqlite3")
    df1 = pd.read_sql_query(query, conn)

    if dict1[year] is None:
        dict1[year] = df1
    
    conn.close()
    print(f"Done with {year}. Number of rows : {len(dict1[year])}")


count_list = []
def start_process(year, query, dict1):
    p = Process(target=run_query, args=(query, year, dict1))
    p.start()
    return p
start_time = time.time()

processes = []

def create_exec(supplier, importer, product, start, end):
    lis = [supplier[:-4], importer[:-4], product[:-4]]
    string = ""
    for i in lis:
        if i.strip() != "":
            string += i + " and "
    string = string[:-5]
    print(string)
    # cols = "BE_NO, STRFTIME('%d-%m-%Y', BEDATE) as date, HS_CODE, PRODUCT_DESCRIPTION, QUANTITY, UNIT, ASSESS_VALUE_INR, UNIT_PRICE_INR, ASSESS_VALUE_USD, UNIT_PRICE_USD, TOTAL_DUTY, TOTAL_DUTY_BE_WISE, APPLICABLE_DUTY_INR, EXCHANGE_RATE_USD, ITEM_RATE_INV_CURR, VALUE_INV_CURR, INVOICE_CURRENCY, ASSESS_GROUP, IMPORTER_CODE, IMPORTER_NAME, IMPORTER_ADDRESS, IMPORTER_CITY, IMPORTER_PIN, IMPORTER_STATE, SUPPLIER_CODE, SUPPLIER_NAME, SUPPLIER_ADDRESS, SUPPLIER_COUNTRY, FOREIGN_PORT, FOREIGN_COUNTRY, FOREIGN_REGIONS, CHA_NAME, CHA_PAN, IEC, IEC_CODE, INVOICE_NUMBER, INVOICE_SR_NO, ITEM_NUMBER, HSCODE_2DIGIT, HSCODE_4DIGIT, TYPE, INDIAN_PORT, SHIPMENT_MODE, INDIAN_REGIONS, SHIPMENT_PORT, HSCODE_6DIGIT, BCD_NOTN, BCD_RATE, BCD_AMOUNT_INR, CVD_NOTN, CVD_RATE, CVD_AMOUNT_INR, IGST_AMOUNT_INR, GST_CESS_AMOUNT_INR, REMARK, INCOTERMS, TOTAL_FREIGHT_VALUE_FORGN_CUR, FREIGHT_CURRENCY, TOTAL_INSU_VALUE_FORGN_CUR, INSURANCE_CURRENCY, TOTAL_INVOICE_VALUE_INR, INSURANCE_VALUE_INR, TOTAL_GROSS_WEIGHT, TOTAL_FREIGHT_VALUE_INR, GROSS_WEIGHT_UNIT, CUSTOM_NOTIFICATION, STANDARD_QUANTITY, STANDARD_QUANTITY_UNIT"
    # cols = "STRFTIME('%d-%m-%Y', BEDATE) as date"
    cols = "*"
    year = start[:4]

    if start[5:] == "01-01" and end[5:] == "12-31":
        bedate = ""
    else:
        bedate = "and BEDATE between '{}' and '{}'".format(start, end)

    exec_str = "select {} from Data_{} where ROWID in ( select ROWID from Data_{}_virt_searcher where {}) {}".format(cols, year, year, string, bedate)
 
    return exec_str


def create_batch(start_date, end_date, supplier, importer, product):

    dict1 = Manager().dict()
    for j in range(int(start_date[:4]), int(end_date[:4])+1):
        df_dict[str(j)] = None
        dict1[str(j)] = None
        process_dict[str(j)] = []

    
    start_time_query = time.time()
    for year in df_dict.keys():
        if(year == start_date[:4]):
            if(year == end_date[:4]):
                end = end_date
            else:
                end = str(year)+"-12-31"
            start = start_date

        elif(year == end_date[:4]):
            start = str(year)+"-01-01"
            end = end_date
        else:
            start = str(year)+"-01-01"
            end = str(year)+"-12-31"
        print(start, end)
        query = create_exec(supplier, importer, product, start, end)
        
        process_dict[year].append(start_process(year, query, dict1))
        print(query)
        print()
    
    # flag = True
    # while(flag):
    #     flag = False
    #     for year in df_dict.keys():
    #         if(dict1[year] is not None and df_dict[year] is None):
    #             print("Done with", year)
    #             print(len(dict1[year]))
    #             df_dict[year] = dict1[year]
    #     for year in df_dict.keys():
    #         if(df_dict[year] is None):
    #             flag = True
    #             break    
    for process in process_dict.values():
        for p in process:
            p.join()

    print("Time taken to retrieve results : ", time.time() - start_time_query)
    return dict1

def home(request):
    return render(request, 'SearchApp/Home.html')

def insert(request):
    import create_table

    if request.method == 'POST' and request.FILES.get('file'):
        date = request.POST.get('Month')
        year = date[:4]

        uploaded_file = request.FILES['file']

        print("Started insert operation for : ", request.POST.get('Month'))
        fs = FileSystemStorage(location=tempfile.gettempdir())
        filename = fs.save(uploaded_file.name, uploaded_file)
        file_path = fs.path(filename)
        print("Uploaded file : ", file_path)

        # file_path = easygui.fileopenbox(filetypes=["*.zip"])
        if file_path is None:
            return render(request, 'SearchApp/Insert.html', {"month" : date})
        
        os.makedirs(f"Data/Excel_Files/{year}/{date}", exist_ok=True)
        current_files = os.listdir(f"Data/Excel_Files/{year}/{date}")
        current_files.sort()
        if(len(current_files) > 0):
            current_num = int(current_files[-1].split("_")[-1].split(".")[0])
            current_num += 1
        else:
            current_num = 0

        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            zip_ref.extractall(tempfile.gettempdir())
            # print(os.walk(tempfile.gettempdir()))
            for root, dirs, files in os.walk(tempfile.gettempdir()):
                for file in files:
                    if file.endswith(".xlsx"):
                        shutil.move(os.path.join(root, file), f"Data/Excel_Files/{year}/{date}/{date}_{current_num}.xlsx")
                        print(f"Data/Excel_Files/{year}/{date}/{date}_{current_num}.xlsx")
                        current_num += 1
        task = ProcessingTask.objects.create(status='pending')

        os.remove(file_path)         
        # create_table.check_new_file(task.id, schedule=5)
        p = Process(target=create_table.check_new_file, args=(task.id,))
        p.start()

        return redirect('loading', task_id=task.id)
    
    context = {"month" : datetime.now().strftime("%Y-%m")}
    return render(request, 'SearchApp/Insert.html', context)

def upload(request):
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        
        print(request.POST.get('Month'))
        fs = FileSystemStorage(location=tempfile.gettempdir())
        filename = fs.save(uploaded_file.name, uploaded_file)
        file_path = fs.path(filename)
        print(file_path)

    return render(request, 'SearchApp/Upload.html')

def search_bom(request):
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        
        start_time = time.time()
        print("Starting BOM search")
        fs = FileSystemStorage(location=tempfile.gettempdir())
        filename = fs.save(uploaded_file.name, uploaded_file)
        file_path = fs.path(filename)
        print("Uploaded file : ", file_path)

        df = pd.read_excel(file_path, sheet_name="E_BOM", skiprows=5, skipfooter=2)
        parts = df["MPN"].tolist()

        
        bom_dict = {}
        tempdir = tempfile.gettempdir()+"/BOM"

        if os.path.exists(tempdir):
            shutil.rmtree(tempdir)
        os.mkdir(tempdir)

        
        for part in parts:
            final_df = pd.DataFrame()
            if type(part) == str and part.strip != "":
                part_dict = create_batch("2018-01-01", get_latest_date(), " and", " and", "PRODUCT_DESCRIPTION MATCH '\"" + part + "\"' and")
                for year in part_dict.keys():
                    final_df = pd.concat([final_df, part_dict[year]])
                
                final_df.sort_values(by=["UNIT_PRICE_USD"], inplace=True)
                final_df.to_csv(f"{tempdir}/{part}.csv", index=False)
                bom_dict[part] = f"{part}.csv"
        
        shutil.copy(file_path, f"Data/Results/{uploaded_file.name[:-5]}_pricing.xlsx")
        os.remove(file_path)
        file_path = f"Data/Results/{uploaded_file.name[:-5]}_pricing.xlsx"

        workbook = openpyxl.load_workbook(file_path)
        bom_sheet = workbook["E_BOM"]

        for part_name in bom_dict.keys():        
            csvfile = f"{tempdir}/{part_name}.csv"
            
            worksheet = workbook.create_sheet(part_name)
            with open(csvfile, 'rt', encoding='utf8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        worksheet.cell(row=r + 1, column=c + 1, value=col)
            
            row = df.index[df["MPN"] == part_name].tolist()[0] + 7
            col = df.columns.get_loc("MPN") + 1  # openpyxl uses 1-based indexing

            bom_sheet.cell(row=row, column=col).value = part_name
            bom_sheet.cell(row=row, column=col).hyperlink = f'#{part_name}!A1'
            bom_sheet.cell(row=row, column=col).style = "Hyperlink"            
                        
            for col in worksheet.columns:
                max_length = 0
                col_letter = col[0].column_letter  # Get the column letter (A, B, C, etc.)
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                # Adjust the column width slightly (add padding)
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[col_letter].width = adjusted_width

            worksheet.auto_filter.ref = 'A1:BP1'
            worksheet.freeze_panes = worksheet['A2']  # Freezes row 1
            
        workbook.save(file_path)
        workbook.close()
        
        print("Time taken for all parts : ", time.time() - start_time)

        with open(file_path, "rb") as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{uploaded_file.name[:-5]}_pricing.xlsx"'
            return response
        
    return render(request, 'SearchApp/Search_BOM.html')

def search(request):
    start_time = time.time()
    global dl

    latest_date = get_latest_date()

    if request.method == "POST":
        start_date = request.POST.get('from_date').upper()
        end_date = request.POST.get('to_date').upper()
        supplier = request.POST.get('SN').upper()
        importer = request.POST.get('IN').upper()
        product = request.POST.get('PD').upper()

        
        df = pd.DataFrame()
        query_type = ""

        if supplier == "":
            supplier = " and"
            query_type += "0"
        else:
            supplier = "SUPPLIER_NAME MATCH '" + supplier +"' and"
            query_type += "1"

        if importer == "":
            importer = " and"
            query_type += "0"
        else:
            importer = "IMPORTER_NAME MATCH '" + importer + "' and"
            query_type += "1"

        if product == "":
            product = " and"
            query_type += "0"
        else:
            product = "PRODUCT_DESCRIPTION MATCH '\"" + product + "\"' and"
            query_type += "1"


        if start_date == "" or int(start_date[:4]) < 2018:
            start_date = "2018-01-01"

        
        if end_date == "" or \
            datetime.strptime(end_date, '%Y-%m-%d').date() < datetime.strptime(start_date, '%Y-%m-%d').date() or \
                datetime.strptime(end_date, '%Y-%m-%d').date() > datetime.strptime(latest_date, '%Y-%m-%d').date():
            
            end_date = latest_date


        result_name = start_date+"_"+(("S-"+supplier[21:-5]+"_"+"I-"+importer[21:-5]+"_"+"P-"+product[27:-5].strip("\"")).rstrip("_").lstrip("_"))+"_"+end_date
        
        if os.path.exists("Data/Results/"+result_name+"/"):
            print("Results folder already exists")
            while(not os.path.exists(f"Data/Results/{result_name}/{result_name}.xlsx")):
                pass
            print("File exists")
            
        else:
            print("Results folder does not exist, creating now")
            os.mkdir("Data/Results/"+result_name+"/")

            dict1 = create_batch(start_date, end_date, supplier, importer, product)

            for year in dict1.keys():
                df = pd.concat([df, dict1[year]])

            print(df.shape)
            df.to_csv("Data/Results/"+result_name+"/"+f"{result_name}.csv", index=False)

            csvfile = f"Data/Results/{result_name}/{result_name}.csv"
            workbook = Workbook(f"Data/Results/{result_name}/{result_name}.xlsx")
            worksheet = workbook.add_worksheet()
            with open(csvfile, 'rt', encoding='utf8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        worksheet.write(r, c, col)
            worksheet.autofit()
            worksheet.autofilter('A1:BP1')
            worksheet.freeze_panes(1, 0)
            workbook.close()

        print("Time taken : ",time.time() - start_time)
        
        with open("Data/Results/"+result_name+"/"+f"{result_name}.xlsx", "rb") as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{result_name}.xlsx"'
            return response
    
    
    
    #send context of last month date
    context = {"last_month" : latest_date}
    return render(request, 'SearchApp/Search-page.html', context)

def loading(request, task_id):
    # task = ProcessingTask.objects.get(id=task_id)
    # Pass the task's status to the template to display progress
    return render(request, 'SearchApp/Loading.html', {'task_id': task_id})

def progress_status(request, task_id):
    task = ProcessingTask.objects.get(id=task_id)
    return JsonResponse({'progress': task.progress})